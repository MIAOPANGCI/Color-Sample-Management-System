"""
封样件及色板接收登记管理系统 - Flask主应用
技术栈: Python Flask + SQLite + JWT + openpyxl
"""
import sqlite3
import hashlib
import json
import jwt
import math
import random
import string
from datetime import datetime, date, timedelta, timezone
from functools import wraps
from flask import Flask, request, jsonify, g, send_file, render_template, send_from_directory
import os
from flask_cors import CORS
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
CORS(app)
app.config['SECRET_KEY'] = 'seal_color_system_secret_key_2024'
app.config['DB_PATH'] = 'seal_samples.db'
app.config['JWT_EXPIRATION_HOURS'] = 24

TZ_OFFSET = timezone(timedelta(hours=8))

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DB_PATH'])
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def get_china_time():
    return datetime.now(TZ_OFFSET).strftime('%Y-%m-%d %H:%M:%S')

def get_china_date():
    return datetime.now(TZ_OFFSET).strftime('%Y-%m-%d')

# ==================== 认证装饰器 ====================

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # 支持Header和Query两种方式传递token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            token = request.args.get('token', '')
        if not token:
            return jsonify({'success': False, 'message': '未提供认证Token'}), 401
        try:
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = dict(get_user_by_id(data['user_id']))
            if not current_user:
                return jsonify({'success': False, 'message': '用户不存在'}), 401
            # 检查账户是否被禁用/删除
            if not current_user.get('is_active', 1):
                return jsonify({'success': False, 'message': '账户已被停用，请联系管理员', 'code': 'ACCOUNT_DISABLED'}), 401
            g.current_user = current_user
            # 更新最后活跃时间（用于在线状态）
            db = get_db()
            db.execute("UPDATE users SET last_active=CURRENT_TIMESTAMP WHERE id=?", (current_user['id'],))
            db.commit()
        except jwt.ExpiredSignatureError:
            return jsonify({'success': False, 'message': 'Token已过期'}), 401
        except Exception as e:
            return jsonify({'success': False, 'message': f'Token无效: {str(e)}'}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not g.current_user.get('is_admin'):
            return jsonify({'success': False, 'message': '需要管理员权限'}), 403
        return f(*args, **kwargs)
    return decorated

# ==================== 数据库初始化 ====================

def init_db():
    """创建10张表 + 预置admin用户 + 预置ΔE阈值设置"""
    conn = sqlite3.connect(app.config['DB_PATH'])
    cursor = conn.cursor()
    
    # 表1: users 用户表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            real_name TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 1,
            must_change_password INTEGER DEFAULT 0,
            invitation_code_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (invitation_code_id) REFERENCES seal_invitation_codes(id)
        )
    ''')

    # 兼容：补充 last_active 字段
    try: cursor.execute("ALTER TABLE users ADD COLUMN last_active TIMESTAMP")
    except: pass

    # 表2: seal_invitation_codes 邀请码表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_invitation_codes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            note TEXT,
            max_uses INTEGER DEFAULT 1,
            used_count INTEGER DEFAULT 0,
            expires_at DATE,
            is_active INTEGER DEFAULT 1,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES users(id)
        )
    ''')
    
    # 表3: seal_samples 封样件台账表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            序号 INTEGER UNIQUE,
            项目 TEXT,
            封样件名称 TEXT,
            签署人 TEXT,
            签署人日期 DATE,
            有效期 DATE,
            状态 TEXT DEFAULT 'normal',
            备注 TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 兼容旧数据库：seal_samples 补充提醒天数列
    try:
        cursor.execute("ALTER TABLE seal_samples ADD COLUMN 提醒天数 INTEGER DEFAULT 30")
    except:
        pass

    # 表4: seal_color_samples 色板台账表(29字段)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_color_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            序号 INTEGER,
            客户 TEXT,
            适用车型 TEXT,
            颜色名称 TEXT,
            样板供应商 TEXT,
            颜色色值转化码 TEXT,
            纹理代码 TEXT,
            光泽度 TEXT,
            供应商代码 TEXT,
            制作信息 TEXT,
            接收数量 INTEGER,
            当前持有数量 INTEGER DEFAULT 0,
            接收日期 DATE,
            使用的光源角度 TEXT,
            L值 REAL,
            a值 REAL,
            b值 REAL,
            c值 REAL,
            h值 REAL,
            ΔL值 REAL,
            Δa值 REAL,
            Δb值 REAL,
            Δc值 REAL,
            Δh值 REAL,
            ΔE值 REAL,
            备注 TEXT,
            有效期 DATE,
            状态 TEXT DEFAULT 'normal',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 兼容旧数据库：补充可能缺失的列
    missing_cols_color = {
        '备注': 'TEXT',
        '有效期': 'DATE',
        'ΔL值': 'REAL', 'Δa值': 'REAL', 'Δb值': 'REAL', 'Δc值': 'REAL', 'Δh值': 'REAL', 'ΔE值': 'REAL',
        '提醒天数': 'INTEGER DEFAULT 30'
    }
    for col, col_type in missing_cols_color.items():
        try:
            cursor.execute(f"ALTER TABLE seal_color_samples ADD COLUMN {col} {col_type}")
        except:
            pass  # 列已存在则忽略

    # 表5: seal_send_records 寄出台账表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_send_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sample_id INTEGER,
            客户 TEXT,
            颜色名称 TEXT,
            对方单位 TEXT,
            寄出数量 INTEGER,
            寄出日期 DATE,
            经手人 TEXT,
            备注 TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sample_id) REFERENCES seal_color_samples(id)
        )
    ''')
    
    # 表6: seal_expiry_management 有效期管理表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_expiry_management (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_type TEXT,
            item_id INTEGER,
            有效期类型 TEXT,
            有效期时长 INTEGER,
            有效期单位 TEXT,
            有效期截止日期 DATE,
            提醒天数 INTEGER DEFAULT 30,
            备注 TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 表7: seal_evaluation_records 评定记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_evaluation_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_type TEXT,
            item_id INTEGER,
            评定结果 TEXT,
            评定人 TEXT,
            评定日期 DATE,
            当前L值 REAL,
            当前a值 REAL,
            当前b值 REAL,
            计算ΔE值 REAL,
            评定说明 TEXT,
            新有效期截止日 DATE,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # 表8: seal_scrapped_samples 报废记录表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_scrapped_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_type TEXT,
            item_id INTEGER,
            报废原因 TEXT,
            报废类型 TEXT,
            报废日期 DATE,
            报废审批人 TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            备注 TEXT
        )
    ''')
    
    # 表9: seal_system_settings 系统设置表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_system_settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key TEXT UNIQUE NOT NULL,
            value TEXT NOT NULL,
            description TEXT,
            updated_by INTEGER,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (updated_by) REFERENCES users(id)
        )
    ''')
    
    # 预置admin用户
    existing_admin = cursor.execute("SELECT id FROM users WHERE username='admin'").fetchone()
    if not existing_admin:
        pwd_hash = hashlib.sha256('admin'.encode()).hexdigest()
        cursor.execute(
            "INSERT INTO users (username, password_hash, real_name, is_admin, must_change_password) VALUES (?, ?, ?, ?, ?)",
            ('admin', pwd_hash, '系统管理员', 1, 1)
        )
    
    # 预置ΔE阈值设置
    default_settings = [
        ('delta_e_excellent', '1.0', 'ΔE优秀阈值上限(低于此值显示绿色优秀)'),
        ('delta_e_good', '2.0', 'ΔE合格阈值上限(介于优秀和此值间显示黄色合格)'),
        ('delta_e_warning', '999.0', 'ΔE需关注阈值(超过此值显示红色需关注)')
    ]
    for key, value, desc in default_settings:
        exists = cursor.execute("SELECT id FROM seal_system_settings WHERE key=?", (key,)).fetchone()
        if not exists:
            cursor.execute(
                "INSERT INTO seal_system_settings (key, value, description) VALUES (?, ?, ?)",
                (key, value, desc)
            )

    # 表10: seal_user_table_configs 用户表格配置表(筛选条件+列显示)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS seal_user_table_configs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            page_key TEXT NOT NULL,
            config_type TEXT NOT NULL,
            config_data TEXT NOT NULL DEFAULT '{}',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, page_key, config_type),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    ''')

    conn.commit()

    # 兼容：为处置记录表添加软删除字段 + 回滚所需字段
    for _sql in [
        "ALTER TABLE seal_evaluation_records ADD COLUMN is_deleted INTEGER DEFAULT 0",
        "ALTER TABLE seal_evaluation_records ADD COLUMN 旧有效期 DATE",
        "ALTER TABLE seal_scrapped_samples ADD COLUMN is_deleted INTEGER DEFAULT 0",
        "ALTER TABLE seal_scrapped_samples ADD COLUMN 旧状态 TEXT"
    ]:
        try: cursor.execute(_sql)
        except: pass
    conn.commit()
    conn.close()

# ==================== 工具函数 ====================

def get_user_by_id(user_id):
    db = get_db()
    return db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()

def calculate_delta_e(baseline_L, baseline_a, baseline_b, current_L, current_a, current_b):
    """CIE76 ΔE色差计算"""
    try:
        return math.sqrt((current_L - baseline_L)**2 + (current_a - baseline_a)**2 + (current_b - baseline_b)**2)
    except:
        return 0.0

def get_expiry_status(expiry_date_str, remind_days=30):
    """判断有效期状态, remind_days为到期前多少天转为待评定"""
    try:
        remind_days = int(remind_days)
    except Exception:
        remind_days = 30
    if not expiry_date_str:
        return 'normal'
    try:
        _s = str(expiry_date_str)
        expiry = datetime.strptime(_s, '%Y-%m-%d').date() if isinstance(expiry_date_str, str) else expiry_date_str
        days_left = (expiry - date.today()).days
        if days_left > remind_days:
            return 'normal'
        elif days_left > 0:
            return 'pending_eval'
        else:
            return 'expired'
    except Exception as ex:
        app.logger.warning(f'[get_expiry] EXCEPTION: {type(ex).__name__}: {ex}')
        return 'normal'

def generate_invitation_code():
    """生成16位随机邀请码"""
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=16))

def paginate(query, page, page_size):
    """分页处理，page_size<=0时返回全量"""
    total = len(query)
    if page_size <= 0:
        return {
            'items': [dict(item) for item in query],
            'total': total,
            'page': 1,
            'page_size': total,
            'total_pages': 1
        }
    offset = (page - 1) * page_size
    items = query[offset:offset + page_size]
    return {
        'items': [dict(item) for item in items],
        'total': total,
        'page': page,
        'page_size': page_size,
        'total_pages': (total + page_size - 1) // page_size
    }

def row_to_dict(row):
    if row:
        return dict(row)
    return None

def apply_dynamic_filters(sql, params, request_args):
    """处理前端传来的动态筛选条件 (f_field_i/f_op_i/f_val_i)"""
    i = 0
    while f'f_field_{i}' in request_args:
        field = request_args[f'f_field_{i}']
        op = request_args.get(f'f_op_{i}', 'contains')
        value = request_args.get(f'f_val_{i}', '')
        
        if not field or not value or not op:
            i += 1
            continue
        
        # 防止SQL注入：只允许字段名中的中文字母数字和下划线
        safe_field = ''.join(c for c in field if c.isalnum() or c in ('_', 'Δ', '-'))
        if not safe_field:
            i += 1
            continue
        
        # 操作符映射到SQL
        op_sql_map = {
            'contains': f" LIKE ?",
            'not_contains': " NOT LIKE ?",
            'equals': " = ?",
            'not_equals': " != ?",
            'gt': " > ?",
            'gte': " >= ?",
            'lt': " < ?",
            'lte': " <= ?",
            'before': " < ?",
            'after': " > ?",
            'is_empty': " IS NULL OR ''"
        }
        
        if op == 'is_empty':
            sql += f" AND ({safe_field} IS NULL OR {safe_field} = '')"
        else:
            sql_op = op_sql_map.get(op, ' LIKE ?')
            if 'LIKE' in sql_op:
                sql += f" AND {safe_field}{sql_op}"
                params.append('%' + value + '%')
            else:
                sql += f" AND {safe_field}{sql_op}"
                params.append(value)
        
        i += 1
    
    return sql, params


# ==================== API: 认证模块 ====================

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user:
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    user_dict = dict(user)
    if not user_dict.get('is_active'):
        return jsonify({'success': False, 'message': '账户已被禁用，请联系管理员'}), 403
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    if pwd_hash != user_dict.get('password_hash', ''):
        return jsonify({'success': False, 'message': '用户名或密码错误'}), 401
    
    token_payload = {
        'user_id': user_dict['id'],
        'username': user_dict['username'],
        'is_admin': user_dict['is_admin'],
        'must_change_password': user_dict.get('must_change_password', 0)
    }
    token = jwt.encode(token_payload, app.config['SECRET_KEY'], algorithm='HS256')
    # 登录成功立即更新在线状态
    db.execute("UPDATE users SET last_active=CURRENT_TIMESTAMP WHERE id=?", (user_dict['id'],))
    db.commit()
    return jsonify({
        'success': True,
        'message': '登录成功',
        'data': {
            'token': token,
            'user': {
                'id': user_dict['id'],
                'username': user_dict['username'],
                'real_name': user_dict['real_name'],
                'is_admin': user_dict['is_admin'],
                'must_change_password': user_dict.get('must_change_password', 0)
            }
        }
    })

@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data.get('username', '')
    real_name = data.get('real_name', '')
    password = data.get('password', '')
    confirm_pwd = data.get('confirmPassword', '') or data.get('confirm_password', '')
    invite_code = data.get('invitationCode', '') or data.get('invitation_code', '')
    
    db = get_db()
    
    if not all([username, real_name, password, confirm_pwd, invite_code]):
        return jsonify({'success': False, 'message': '请填写完整信息'}), 400
    if password != confirm_pwd:
        return jsonify({'success': False, 'message': '两次输入的密码不一致'}), 400
    if len(password) < 6:
        return jsonify({'success': False, 'message': '密码长度不能少于6位'}), 400
    
    # 检查用户名是否已存在
    if db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        return jsonify({'success': False, 'message': '用户名已被注册'}), 400
    
    # 验证邀请码
    code_row = db.execute("SELECT * FROM seal_invitation_codes WHERE code=?", (invite_code,)).fetchone()
    if not code_row:
        return jsonify({'success': False, 'message': '无效的邀请码'}), 400
    code_dict = dict(code_row)
    if not code_dict.get('is_active'):
        return jsonify({'success': False, 'message': '该邀请码已停用'}), 400
    if code_dict.get('expires_at') and str(code_dict['expires_at']) < get_china_date():
        return jsonify({'success': False, 'message': '该邀请码已过期'}), 400
    if code_dict.get('used_count', 0) >= code_dict.get('max_uses', 1):
        return jsonify({'success': False, 'message': '该邀请码使用次数已达上限'}), 400
    
    # 创建用户
    pwd_hash = hashlib.sha256(password.encode()).hexdigest()
    cursor = db.cursor()
    cursor.execute(
        "INSERT INTO users (username, password_hash, real_name, invitation_code_id) VALUES (?, ?, ?, ?)",
        (username, pwd_hash, real_name, code_dict['id'])
    )
    user_id = cursor.lastrowid
    
    # 更新邀请码使用次数
    cursor.execute("UPDATE seal_invitation_codes SET used_count=used_count+1 WHERE id=?", (code_dict['id'],))
    db.commit()
    
    return jsonify({'success': True, 'message': '注册成功，请登录'})

@app.route('/api/auth/change-password', methods=['PUT'])
@token_required
def change_password():
    data = request.get_json()
    old_pwd = data.get('oldPassword', '')
    new_pwd = data.get('newPassword', '')
    confirm_pwd = data.get('confirmPassword', '')
    
    if not all([old_pwd, new_pwd, confirm_pwd]):
        return jsonify({'success': False, 'message': '请填写完整信息'}), 400
    if new_pwd != confirm_pwd:
        return jsonify({'success': False, 'message': '两次输入的新密码不一致'}), 400
    if len(new_pwd) < 6:
        return jsonify({'success': False, 'message': '新密码长度不能少于6位'}), 400
    
    old_hash = hashlib.sha256(old_pwd.encode()).hexdigest()
    db = get_db()
    user = db.execute("SELECT password_hash FROM users WHERE id=?", (g.current_user['id'],)).fetchone()
    if dict(user)['password_hash'] != old_hash:
        return jsonify({'success': False, 'message': '原密码不正确'}), 400
    
    new_hash = hashlib.sha256(new_pwd.encode()).hexdigest()
    db.execute("UPDATE users SET password_hash=?, must_change_password=0 WHERE id=?", 
               (new_hash, g.current_user['id']))
    db.commit()
    
    return jsonify({'success': True, 'message': '密码修改成功，请重新登录'})

@app.route('/api/auth/ping', methods=['GET','POST'])
@token_required
def ping():
    """心跳接口 - 更新用户在线时间"""
    db = get_db()
    db.execute("UPDATE users SET last_active=CURRENT_TIMESTAMP WHERE id=?", (g.current_user['id'],))
    db.commit()
    return jsonify({'success': True})

@app.route('/api/auth/logout', methods=['POST'])
@token_required
def logout():
    """退出登录 - 清除在线状态"""
    db = get_db()
    db.execute("UPDATE users SET last_active=NULL WHERE id=?", (g.current_user['id'],))
    db.commit()
    return jsonify({'success': True, 'message': '已退出登录'})


# ==================== API: 封样件 CRUD ====================

@app.route('/api/seal-samples', methods=['GET'])
@token_required
def list_seal_samples():
    db = get_db()
    page = int(request.args.get('page', 1))
    _ps = request.args.get('pageSize')
    page_size = int(_ps) if _ps and _ps.strip() else 0  # 不传则全量
    
    sql = "SELECT * FROM seal_samples WHERE 1=1"
    params = []
    
    # 动态筛选
    sql, params = apply_dynamic_filters(sql, params, request.args)
    
    # 兼容旧参数
    search = request.args.get('search', '')
    project_filter = request.args.get('project', '')
    if search:
        sql += " AND (封样件名称 LIKE ? OR 签署人 LIKE ?)"
        params.extend(['%' + search + '%'] * 2)
    if project_filter:
        sql += " AND 项目 LIKE ?"
        params.append('%' + project_filter + '%')
    
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    result = paginate(rows, page, page_size)
    for item in result['items']:
        if item['状态'] == 'normal':
            item['状态'] = get_expiry_status(item['有效期'], item.get('提醒天数', 30))
    return jsonify({'success': True, 'data': result})

@app.route('/api/seal-samples', methods=['POST'])
@token_required
def create_seal_sample():
    data = request.get_json()
    required_fields = ['项目', '封样件名称', '签署人', '签署人日期', '有效期']
    for f in required_fields:
        if f not in data or not str(data[f]).strip():
            return jsonify({'success': False, 'message': f'{f}为必填项'}), 400
    
    db = get_db()
    cursor = db.cursor()
    # 自动生成序号：GKYJ-年月日时分秒
    data['序号'] = f"GKYJ-{datetime.now(TZ_OFFSET).strftime('%Y%m%d%H%M%S')}"
    
    cursor.execute('''INSERT INTO seal_samples (序号, 项目, 封样件名称, 签署人, 签署人日期, 有效期, 提醒天数, 备注, 状态) 
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                   (data['序号'], data['项目'], data['封样件名称'], data['签署人'], 
                    data['签署人日期'], data['有效期'], data.get('提醒天数', 30), data.get('备注', ''),
                    get_expiry_status(data['有效期'], data.get('提醒天数', 30))))
    db.commit()
    return jsonify({'success': True, 'message': '封样件添加成功', 'data': {'id': cursor.lastrowid}})

@app.route('/api/seal-samples/<int:id>', methods=['GET'])
@token_required
def get_seal_sample(id):
    db = get_db()
    sample = db.execute("SELECT * FROM seal_samples WHERE id=?", (id,)).fetchone()
    if not sample:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    data = dict(sample)
    # 动态计算状态（与列表接口一致）
    if data.get('状态') == 'normal':
        data['状态'] = get_expiry_status(data.get('有效期'), data.get('提醒天数', 30))
    return jsonify({'success': True, 'data': data})

@app.route('/api/seal-samples/<int:id>', methods=['PUT'])
@token_required
def update_seal_sample(id):
    data = request.get_json()
    db = get_db()
    sample = db.execute("SELECT * FROM seal_samples WHERE id=? AND 状态!='scrapped'", (id,)).fetchone()
    if not sample:
        return jsonify({'success': False, 'message': '记录不存在或已报废'}), 404
    
    cursor = db.cursor()
    new_expiry = data.get('有效期', sample['有效期'])
    new_remind = data.get('提醒天数', 30)
    cursor.execute('''UPDATE seal_samples SET 序号=?, 项目=?, 封样件名称=?, 签署人=?, 
                      签署人日期=?, 有效期=?, 提醒天数=?, 备注=?, 状态=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                   (data.get('序号', sample['序号']), data.get('项目', sample['项目']),
                    data.get('封样件名称', sample['封样件名称']), data.get('签署人', sample['签署人']),
                    data.get('签署人日期', sample['签署人日期']), new_expiry,
                    new_remind, data.get('备注', sample['备注']),
                    get_expiry_status(new_expiry, int(new_remind) if new_remind else 30), id))
    db.commit()
    return jsonify({'success': True, 'message': '更新成功'})

@app.route('/api/seal-samples/<int:id>', methods=['DELETE'])
@token_required
def delete_seal_sample(id):
    db = get_db()
    if not db.execute("SELECT id FROM seal_samples WHERE id=?", (id,)).fetchone():
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    db.execute("DELETE FROM seal_samples WHERE id=?", (id,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功'})

@app.route('/api/seal-samples/export', methods=['GET'])
@token_required
def export_seal_samples():
    db = get_db()
    rows = db.execute("SELECT * FROM seal_samples ORDER BY id DESC").fetchall()
    # 状态映射为中文
    status_map = {'normal': '正常', 'pending_eval': '待评定', 'expired': '已过期', 'scrapped': '已报废'}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "封样件台账"
    headers = ['ID','序号','项目','封样件名称','签署人','签署人日期','有效期','提醒天数','状态','备注','创建时间']
    ws.append(headers)
    for row in rows:
        d = dict(row)
        st = d.get('状态', '')
        st_cn = status_map.get(st, st)
        if st == 'normal':
            st_cn = get_expiry_status(d.get('有效期'), d.get('提醒天数', 30))  # 动态计算
            st_cn = status_map.get(st_cn, st_cn)
        ws.append([d['id'], d['序号'], d['项目'], d['封样件名称'], d['签署人'],
                   d['签署人日期'], d['有效期'], d.get('提醒天数', 30), st_cn, d['备注'], d['created_at']])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='封样件台账.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/seal-samples/import', methods=['POST'])
@token_required
def import_seal_samples():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '请选择要导入的文件'}), 400
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 检测表头是否有ID列（从导出文件重新导入时会带ID），有则跳过
        _header_row = [cell.value for cell in ws[1]]
        _skip_first = len(_header_row) > 0 and str(_header_row[0]).strip().upper() == 'ID'

        rows = []
        for row in ws.iter_rows(min_row=2):
            cells = list(row)
            if _skip_first:
                cells = cells[1:]  # 跳过ID列
            row_vals = []
            for cell in cells:
                v = cell.value
                if v is None:
                    row_vals.append('')
                elif isinstance(v, (int, float, str, bool)):
                    row_vals.append(str(v).strip())
                else:
                    try:
                        row_vals.append(str(v).strip())
                    except Exception:
                        row_vals.append('')
            rows.append(row_vals)

        db = get_db()
        count = 0
        errors = []
        # 与导出模板一致：序号,项目,封样件名称,签署人,签署人日期,有效期,提醒天数,备注（状态默认normal，创建时间自动生成）
        field_names = ['序号','项目','封样件名称','签署人','签署人日期','有效期','提醒天数','备注']
        status_map_cn = {'正常':'normal','待评定':'pending_eval','已过期':'expired','已报废':'scrapped'}
        for idx, r in enumerate(rows):
            if not r or not any(v for v in r): continue
            try:
                vals = [(str(x).strip() if x else '') for x in (r + [''] * len(field_names))][:len(field_names)]
                # 日期字段截取 YYYY-MM-DD
                vals[4] = vals[4][:10] if len(vals[4]) >= 10 else ''
                vals[5] = vals[5][:10] if len(vals[5]) >= 10 else ''

                db.execute("""INSERT INTO seal_samples (序号, 项目, 封样件名称, 签署人, 签署人日期, 有效期, 提醒天数, 备注, 状态)
                             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                          (vals[0], vals[1], vals[2], vals[3], vals[4],
                           vals[5], int(vals[6]) if vals[6].isdigit() else 30,
                           vals[7], 'normal'))
                count += 1
            except Exception as e2:
                safe_row = [str(x)[:20] if x else '' for x in r[:3]]
                errors.append(f"第{idx+2}行({', '.join(safe_row)}): {e2}")
        db.commit()
        # 自动重算：根据有效期+提醒天数计算真实状态
        try:
            new_rows = db.execute(
                "SELECT id, 有效期, COALESCE(提醒天数,30) FROM seal_samples WHERE id > (SELECT MAX(id) FROM seal_samples) - ?",
                (count,)
            ).fetchall()
            for rid, exp, rd in new_rows:
                real_status = get_expiry_status(exp, int(rd) if rd else 30)
                db.execute("UPDATE seal_samples SET 状态=? WHERE id=?", (real_status, rid))
            db.commit()
        except Exception:
            pass
        msg = f'成功导入{count}条数据'
        if errors:
            msg += f'，{len(errors)}条跳过（首条错误: {errors[0]}）'
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500


# ==================== API: 色板 CRUD ====================

@app.route('/api/color-samples', methods=['GET'])
@token_required
def list_color_samples():
    db = get_db()
    page = int(request.args.get('page', 1))
    _ps = request.args.get('pageSize')
    page_size = int(_ps) if _ps and _ps.strip() else 0  # 不传则全量
    
    sql = "SELECT * FROM seal_color_samples WHERE 1=1"
    params = []
    
    # 动态筛选
    sql, params = apply_dynamic_filters(sql, params, request.args)
    
    # 兼容旧参数
    search = request.args.get('search', '')
    customer = request.args.get('customer', '')
    model = request.args.get('model', '')
    color_name = request.args.get('colorName', '')
    supplier = request.args.get('supplier', '')
    if search:
        sql += " AND (颜色名称 LIKE ? OR 客户 LIKE ? OR 适用车型 LIKE ?)"
        params.extend(['%' + search + '%'] * 3)
    if customer:
        sql += " AND 客户 LIKE ?"
        params.append('%' + customer + '%')
    if model:
        sql += " AND 适用车型 LIKE ?"
        params.append('%' + model + '%')
    if color_name:
        sql += " AND 颜色名称 LIKE ?"
        params.append('%' + color_name + '%')
    if supplier:
        sql += " AND 样板供应商 LIKE ?"
        params.append('%' + supplier + '%')
    
    sql += " ORDER BY id DESC"
    rows = db.execute(sql, params).fetchall()
    result = paginate(rows, page, page_size)
    for item in result['items']:
        if item['状态'] == 'normal':
            item['状态'] = get_expiry_status(item['有效期'], item.get('提醒天数', 30))
    return jsonify({'success': True, 'data': result})

@app.route('/api/color-samples', methods=['POST'])
@token_required
def create_color_sample():
    data = request.get_json()
    required = ['客户', '适用车型', '颜色名称', '接收数量', '有效期']
    for f in required:
        if f not in data or str(data[f]).strip() == '':
            return jsonify({'success': False, 'message': f'{f}为必填项'}), 400
    
    db = get_db()
    # 自动生成序号：GKSB-年月日时分秒
    data['序号'] = f"GKSB-{datetime.now(TZ_OFFSET).strftime('%Y%m%d%H%M%S')}"
    
    fields = ['序号','客户','适用车型','颜色名称','样板供应商','颜色色值转化码','纹理代码','光泽度',
              '供应商代码','制作信息','接收数量','当前持有数量','接收日期','使用的光源角度',
              'L值','a值','b值','c值','h值','ΔL值','Δa值','Δb值','Δc值','Δh值','ΔE值','有效期','提醒天数','备注']
    placeholders = ','.join(['?']*len(fields))
    field_names = ','.join(fields)
    values = [data.get(f, '') for f in fields]
    _calc_status = get_expiry_status(data.get('有效期', ''), data.get('提醒天数', 30))
    values.append(_calc_status)  # 状态
    
    cursor = db.cursor()
    cursor.execute(f"INSERT INTO seal_color_samples ({field_names}, 状态) VALUES ({placeholders}, ?)", values)
    db.commit()
    return jsonify({'success': True, 'message': '色板添加成功', 'data': {'id': cursor.lastrowid}})

@app.route('/api/color-samples/<int:id>', methods=['GET'])
@token_required
def get_color_sample(id):
    db = get_db()
    row = db.execute("SELECT * FROM seal_color_samples WHERE id=?", (id,)).fetchone()
    if not row:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    result = dict(row)
    if result['状态'] == 'normal':
        result['状态'] = get_expiry_status(result['有效期'], result.get('提醒天数', 30))
    return jsonify({'success': True, 'data': result})

@app.route('/api/color-samples/<int:id>', methods=['PUT'])
@token_required
def update_color_sample(id):
    data = request.get_json()
    db = get_db()
    sample = db.execute("SELECT * FROM seal_color_samples WHERE id=? AND 状态!='scrapped'", (id,)).fetchone()
    if not sample:
        return jsonify({'success': False, 'message': '记录不存在或已报废'}), 404
    
    updatable = ['序号','客户','适用车型','颜色名称','样板供应商','颜色色值转化码','纹理代码','光泽度',
                 '供应商代码','制作信息','接收数量','当前持有数量','接收日期','使用的光源角度',
                 'L值','a值','b值','c值','h值','ΔL值','Δa值','Δb值','Δc值','Δh值','ΔE值','有效期','提醒天数','备注']
    set_clause = ','.join([f"{f}=?" for f in updatable])
    values = [data.get(f, sample[f]) for f in updatable]
    new_expiry = data.get('有效期', sample['有效期'])
    new_remind = data.get('提醒天数', sample.get('提醒天数', 30))
    db.execute(f"UPDATE seal_color_samples SET {set_clause}, 状态=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
               values + [get_expiry_status(new_expiry, int(new_remind) if new_remind else 30), id])
    db.commit()
    return jsonify({'success': True, 'message': '更新成功'})

@app.route('/api/color-samples/<int:id>', methods=['DELETE'])
@token_required
def delete_color_sample(id):
    db = get_db()
    if not db.execute("SELECT id FROM seal_color_samples WHERE id=?", (id,)).fetchone():
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    # 检查是否有关联的寄出记录
    send_count = db.execute("SELECT COUNT(*) as cnt FROM seal_send_records WHERE sample_id=?", (id,)).fetchone()['cnt']
    if send_count > 0:
        return jsonify({'success': False, 'message': f'该色板存在{send_count}条寄出记录，不允许删除'}), 400
    db.execute("DELETE FROM seal_color_samples WHERE id=?", (id,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功'})

@app.route('/api/color-samples/export', methods=['GET'])
@token_required
def export_color_samples():
    db = get_db()
    rows = db.execute("SELECT * FROM seal_color_samples ORDER BY id DESC").fetchall()
    
    # 状态映射为中文
    status_map = {'normal': '正常', 'pending_eval': '待评定', 'expired': '已过期', 'scrapped': '已报废'}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "色板台账"
    headers = ['ID','序号','客户','适用车型','颜色名称','样板供应商','颜色色值转化码','纹理代码','光泽度',
               '供应商代码','制作信息','接收数量','当前持有数量','接收日期','使用的光源角度',
               'L值','a值','b值','c值','h值','ΔL值','Δa值','Δb值','Δc值','Δh值','ΔE值','有效期','提醒天数','状态','备注']
    ws.append(headers)
    # 按headers顺序取字段，避免多出created_at/updated_at等列
    db_cols = ['id','序号','客户','适用车型','颜色名称','样板供应商','颜色色值转化码','纹理代码','光泽度',
               '供应商代码','制作信息','接收数量','当前持有数量','接收日期','使用的光源角度',
               'L值','a值','b值','c值','h值','ΔL值','Δa值','Δb值','Δc值','Δh值','ΔE值','有效期','提醒天数','状态','备注']
    for row in rows:
        d = dict(row)
        st = d.get('状态', '')
        st_cn = status_map.get(st, st)
        if st == 'normal':
            st_cn = get_expiry_status(d.get('有效期'))
            st_cn = status_map.get(st_cn, st_cn)
        vals = [d.get(c, '') for c in db_cols]
        vals[-2] = st_cn  # 替换状态为中文
        ws.append(vals)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='色板台账.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/color-samples/import', methods=['POST'])
@token_required
def import_color_samples():
    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'message': '请选择文件'}), 400
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # 检测表头是否有ID列（从导出文件重新导入时会带ID），有则跳过
        _header_row = [cell.value for cell in ws[1]]
        _skip_first = len(_header_row) > 0 and str(_header_row[0]).strip().upper() == 'ID'

        rows = []
        for row in ws.iter_rows(min_row=2):
            cells = list(row)
            if _skip_first:
                cells = cells[1:]  # 跳过ID列
            row_vals = []
            for cell in cells:
                v = cell.value
                if v is None:
                    row_vals.append('')
                elif isinstance(v, (int, float, str, bool)):
                    row_vals.append(str(v).strip())
                else:
                    try:
                        row_vals.append(str(v).strip())
                    except Exception:
                        row_vals.append('')
            rows.append(row_vals)

        db = get_db()
        count = 0
        errors = []
        # 与导出模板一致：全部字段（除ID/创建时间等系统自动生成列）
        field_names = ['序号','客户','适用车型','颜色名称','样板供应商','颜色色值转化码','纹理代码','光泽度',
                       '供应商代码','制作信息','接收数量','当前持有数量','接收日期','使用的光源角度',
                       'L值','a值','b值','c值','h值','ΔL值','Δa值','Δb值','Δc值','Δh值','ΔE值','有效期','提醒天数','状态','备注']
        placeholders = ','.join(['?'] * len(field_names))
        status_map_cn = {'正常':'normal','待评定':'pending_eval','已过期':'expired','已报废':'scrapped'}
        for idx, r in enumerate(rows):
            if not r or not any(v for v in r): continue
            try:
                vals = [(str(x).strip() if x else '') for x in (r + [''] * len(field_names))][:len(field_names)]
                # 状态映射：中文→内部码，空则默认 normal（索引27：有效期→提醒天数→状态）
                status_val = vals[27].strip()
                db_status = status_map_cn.get(status_val, 'normal') if status_val else 'normal'
                vals[27] = db_status

                db.execute(f"INSERT INTO seal_color_samples ({','.join(field_names)}) VALUES ({placeholders})", vals)
                count += 1
            except Exception as e2:
                safe_row = [str(x)[:20] if x else '' for x in r[:3]]
                errors.append(f"第{idx+2}行({', '.join(safe_row)}): {e2}")
        db.commit()
        # 自动重算：ΔE值为空则补算，状态根据有效期+提醒天数自动计算
        try:
            new_rows = db.execute(
                "SELECT id, 有效期, COALESCE(提醒天数,30), COALESCE(`ΔL值`,0), COALESCE(`Δa值`,0), COALESCE(`Δb值`,0), `ΔE值` "
                "FROM seal_color_samples WHERE id > (SELECT MAX(id) FROM seal_color_samples) - ?",
                (count,)
            ).fetchall()
            for rid, exp, rd, dL, da, db_val, de_val in new_rows:
                updates = []
                params = []
                # 状态重算
                real_status = get_expiry_status(exp, int(rd) if rd else 30)
                updates.append("状态=?")
                params.append(real_status)
                # ΔE补算
                if not de_val and (dL or da or db_val):
                    updates.append("`ΔE值`=?")
                    params.append(round(math.sqrt(float(dL)**2 + float(da)**2 + float(db_val)**2), 4))
                if updates:
                    sql = f"UPDATE seal_color_samples SET {','.join(updates)} WHERE id=?"
                    params.append(rid)
                    db.execute(sql, params)
            db.commit()
        except Exception:
            pass
        msg = f'成功导入{count}条数据'
        if errors:
            msg += f'，{len(errors)}条跳过（首条错误: {errors[0]}）'
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500


# ==================== API: 寄出管理 ====================

@app.route('/api/send-records', methods=['GET'])
@token_required
def list_send_records():
    db = get_db()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    
    sql = "SELECT sr.*, cs.客户 as 色板客户, cs.序号 as 序号 FROM seal_send_records sr LEFT JOIN seal_color_samples cs ON sr.sample_id=cs.id WHERE 1=1"
    params = []
    
    # 动态筛选（注意：JOIN查询的字段需要加前缀）
    # 简化处理：对sr表字段和cs表字段分别处理
    i = 0
    while f'f_field_{i}' in request.args:
        field = request.args[f'f_field_{i}']
        op = request.args.get(f'f_op_{i}', 'contains')
        value = request.args.get(f'f_val_{i}', '')
        if not field or not value or not op:
            i += 1; continue
        safe_field = ''.join(c for c in field if c.isalnum() or c in ('_', 'Δ', '-'))
        if not safe_field: i += 1; continue
        
        # 寄出台账的筛选字段可能来自sr表或cs表
        table_prefix = 'sr'
        if safe_field in ('客户','颜色名称'):
            # 这些字段在两个表中都有，优先用sr表的值，但也匹配cs表
            pass
            
        op_sql_map = {
            'contains': ' LIKE ?', 'not_contains': ' NOT LIKE ?',
            'equals': ' = ?', 'not_equals': ' != ?',
            'gt': ' > ?', 'gte': ' >= ?', 'lt': ' < ?', 'lte': ' <= ?'
        }
        sql_op = op_sql_map.get(op, ' LIKE ?')
        
        # 尝试两个字段
        if safe_field == '客户':
            sql += f" AND (sr.客户{sql_op} OR cs.客户{sql_op})"
            val = '%' + value + '%' if 'LIKE' in sql_op else value
            params.extend([val, val])
        elif safe_field == '对方单位':
            sql += f" AND sr.对方单位{sql_op}"
            params.append('%' + value + '%' if 'LIKE' in sql_op else value)
        elif safe_field in ('颜色名称',):
            sql += f" AND sr.颜色名称{sql_op}"
            params.append('%' + value + '%' if 'LIKE' in sql_op else value)
        else:
            sql += f" AND {safe_field}{sql_op}"
            params.append('%' + value + '%' if 'LIKE' in sql_op else value)
        i += 1
    
    # 兼容旧参数
    customer = request.args.get('customer', '')
    recipient = request.args.get('recipient', '')
    if customer:
        sql += " AND (sr.客户 LIKE ? OR cs.客户 LIKE ?)"
        params.extend(['%' + customer + '%'] * 2)
    if recipient:
        sql += " AND sr.对方单位 LIKE ?"
        params.append('%' + recipient + '%')
    
    sql += " ORDER BY sr.id DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify({'success': True, 'data': paginate(rows, page, page_size)})

@app.route('/api/send-records', methods=['POST'])
@token_required
def create_send_record():
    data = request.get_json()
    required = ['sample_id', '对方单位', '寄出数量', '寄出日期']
    for f in required:
        if f not in data or (f == 'sample_id' and not data[f]) or (f != 'sample_id' and not str(data[f]).strip()):
            return jsonify({'success': False, 'message': f'{f}为必填项'}), 400
    
    db = get_db()
    sample = db.execute("SELECT id, 颜色名称, 当前持有数量 FROM seal_color_samples WHERE id=? AND 状态='normal'", 
                        (data['sample_id'],)).fetchone()
    if not sample:
        return jsonify({'success': False, 'message': '关联的色板不存在或已过期，不可操作。'}), 404
    
    sample_dict = dict(sample)
    current_qty = sample_dict['当前持有数量'] or 0
    send_qty = int(data['寄出数量'])
    if send_qty > current_qty:
        return jsonify({'success': False, 'message': f'寄出数量({send_qty})超出当前持有数量({current_qty})'}), 400
    
    cursor = db.cursor()
    cursor.execute("""INSERT INTO seal_send_records (sample_id, 客户, 颜色名称, 对方单位, 寄出数量, 寄出日期, 经手人, 备注)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                   (data['sample_id'], data.get('客户', ''), sample_dict['颜色名称'],
                    data['对方单位'], send_qty, data['寄出日期'], data.get('经手人', ''), data.get('备注', '')))
    # 扣减库存
    cursor.execute("UPDATE seal_color_samples SET 当前持有数量=当前持有数量-? WHERE id=?",
                   (send_qty, data['sample_id']))
    db.commit()
    return jsonify({'success': True, 'message': '寄出记录添加成功，库存已扣减'})

@app.route('/api/send-records/<int:id>', methods=['DELETE'])
@token_required
def delete_send_record(id):
    db = get_db()
    record = db.execute("SELECT * FROM seal_send_records WHERE id=?", (id,)).fetchone()
    if not record:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    rec = dict(record)
    # 恢复库存
    db.execute("UPDATE seal_color_samples SET 当前持有数量=当前持有数量+? WHERE id=?",
               (rec['寄出数量'], rec['sample_id']))
    db.execute("DELETE FROM seal_send_records WHERE id=?", (id,))
    db.commit()
    return jsonify({'success': True, 'message': '删除成功，库存已恢复'})

@app.route('/api/send-records/export', methods=['GET'])
@token_required
def export_send_records():
    db = get_db()
    rows = db.execute("SELECT * FROM seal_send_records ORDER BY id DESC").fetchall()
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "寄出台账"
    ws.append(['ID','色板ID','客户','颜色名称','对方单位','寄出数量','寄出日期','经手人','备注','创建时间'])
    for r in rows:
        d = dict(r); ws.append([d['id'],d['sample_id'],d['客户'],d['颜色名称'],d['对方单位'],
                                d['寄出数量'],d['寄出日期'],d['经手人'],d['备注'],d['created_at']])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='寄出台账.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==================== API: 有效期管理 ====================

@app.route('/api/expiry', methods=['GET'])
@token_required
def list_expiry():
    db = get_db()
    rows = db.execute("SELECT * FROM seal_expiry_management ORDER BY id DESC").fetchall()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})

@app.route('/api/expiry', methods=['POST'])
@token_required
def create_expiry():
    data = request.get_json()
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""INSERT INTO seal_expiry_management (item_type, item_id, 有效期类型, 有效期时长, 有效期单位, 
                      有效期截止日期, 提醒天数, 备注, created_by)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (data.get('item_type'), data.get('item_id'), data.get('有效期类型'),
                    data.get('有效期时长'), data.get('有效期单位'), data.get('有效期截止日期'),
                    data.get('提醒天数', 30), data.get('备注', ''), g.current_user['id']))
    db.commit()
    return jsonify({'success': True, 'message': '有效期规则添加成功'})

@app.route('/api/expiry/<int:id>', methods=['PUT'])
@token_required
def update_expiry(id):
    data = request.get_json()
    db = get_db()
    db.execute("""UPDATE seal_expiry_management SET 有效期类型=?, 有效期时长=?, 有效期单位=?, 
                  有效期截止日期=?, 提醒天数=?, 备注=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
               (data.get('有效期类型'), data.get('有效期时长'), data.get('有效期单位'),
                data.get('有效期截止日期'), data.get('提醒天数'), data.get('备注'), id))
    db.commit()
    return jsonify({'success': True, 'message': '更新成功'})


# ==================== API: 评定提交 ====================

@app.route('/api/evaluations', methods=['GET'])
@token_required
def list_evaluations():
    db = get_db()
    item_type = request.args.get('type', '')
    item_id = request.args.get('itemId', '')
    sql = "SELECT er.* FROM seal_evaluation_records er WHERE 1=1"
    params = []
    if item_type: sql += " AND er.item_type=?"; params.append(item_type)
    if item_id: sql += " AND er.item_id=?"; params.append(item_id)
    sql += " ORDER BY er.id DESC"
    rows = db.execute(sql, params).fetchall()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})

@app.route('/api/evaluations', methods=['POST'])
@token_required
def submit_evaluation():
    data = request.get_json()
    item_type = data.get('item_type', '')
    item_id = data.get('item_id')
    result = data.get('result', '')  # pass/fail
    
    if not all([item_type, item_id, result]):
        return jsonify({'success': False, 'message': '缺少必要参数'}), 400
    
    db = get_db()
    
    # 确定目标表
    table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'
    item = db.execute(f"SELECT * FROM {table} WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    item_dict = dict(item)
    if item_dict.get('状态') == 'scrapped':
        return jsonify({'success': False, 'message': '该项目已报废，无法评定'}), 400
    
    delta_e_val = 0.0
    # 如果是色板评定，自动计算ΔE
    if item_type == 'color' and result == 'pass':
        cur_L = float(data.get('当前L值', 0))
        cur_a = float(data.get('当前a值', 0))
        cur_b = float(data.get('当前b值', 0))
        base_L = float(item_dict.get('L值', 0) or 0)
        base_a = float(item_dict.get('a值', 0) or 0)
        base_b = float(item_dict.get('b值', 0) or 0)
        delta_e_val = round(calculate_delta_e(base_L, base_a, base_b, cur_L, cur_a, cur_b), 4)
    
    cursor = db.cursor()

    # 保存旧有效期（用于删除时回滚）
    _old_expiry = item_dict.get('有效期') or ''

    # 写入评定记录
    cursor.execute("""INSERT INTO seal_evaluation_records (item_type, item_id, 评定结果, 评定人, 评定日期, 
                      当前L值, 当前a值, 当前b值, 计算ΔE值, 评定说明, 新有效期截止日, 旧有效期, created_by)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (item_type, item_id, result, g.current_user['real_name'], get_china_date(),
                    data.get('当前L值'), data.get('当前a值'), data.get('当前b值'),
                    delta_e_val, data.get('评定说明', ''), data.get('新有效期截止日', ''),
                    _old_expiry, g.current_user['id']))
    
    if result == 'pass':
        # 合格续期：更新有效期
        new_exp = data.get('新有效期截止日')
        if new_exp:
            db.execute(f"UPDATE {table} SET 有效期=?, 状态='normal', updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (new_exp, item_id))
    elif result == 'fail':
        # 不合格 → 报废流程由单独的 scrap API处理
        pass
    
    db.commit()
    
    action_msg = '合格续期' if result == 'pass' else '不合格'
    return jsonify({'success': True, 'message': f'评定提交成功({action_msg})', 'data': {'deltaE': delta_e_val}})


# ==================== API: 报废操作 ====================

@app.route('/api/scrap', methods=['POST'])
@token_required
def scrap_item():
    data = request.get_json()
    item_type = data.get('item_type', '')
    item_id = data.get('item_id')
    reason = data.get('报废原因', '')
    scrap_type = data.get('报废类型', '')
    
    if not all([item_type, item_id, reason, scrap_type]):
        return jsonify({'success': False, 'message': '请填写完整的报废信息'}), 400
    
    db = get_db()
    table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'
    item = db.execute(f"SELECT * FROM {table} WHERE id=?", (item_id,)).fetchone()
    if not item:
        return jsonify({'success': False, 'message': '记录不存在'}), 404
    if dict(item).get('状态') == 'scrapped':
        return jsonify({'success': False, 'message': '该项目已报废'}), 400
    
    cursor = db.cursor()

    # 保存旧状态（用于删除时回滚）
    _old_status = dict(item).get('状态') or 'normal'

    # 写入报废记录
    name_field = '封样件名称' if item_type == 'seal' else '颜色名称'
    cursor.execute("""INSERT INTO seal_scrapped_samples (item_type, item_id, 报废原因, 报废类型, 报废日期, 报废审批人, created_by, 备注, 旧状态)
                      VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                   (item_type, item_id, reason, scrap_type, get_china_date(),
                    g.current_user['real_name'], g.current_user['id'], data.get('备注', ''),
                    _old_status))
    # 更新状态为报废
    db.execute(f"UPDATE {table} SET 状态='scrapped', updated_at=CURRENT_TIMESTAMP WHERE id=?", (item_id,))
    db.commit()
    
    return jsonify({'success': True, 'message': '报废操作成功，该记录已锁定'})

@app.route('/api/scrap/<int:id>', methods=['DELETE'])
@token_required
def delete_scrap(id):
    if not g.current_user.get('is_admin'):
        return jsonify({'success': False, 'message': '仅管理员可执行此操作'}), 403
    db = get_db()
    scrap = db.execute("SELECT * FROM seal_scrapped_samples WHERE id=?", (id,)).fetchone()
    if not scrap:
        return jsonify({'success': False, 'message': '报废记录不存在'}), 404
    is_permanent = request.args.get('permanent', '') == '1'
    scrap_dict = dict(scrap)
    item_type = scrap_dict['item_type']
    item_id = scrap_dict['item_id']
    db.execute("DELETE FROM seal_scrapped_samples WHERE id=?", (id,))
    if is_permanent:
        # 永久删除：同时删除原记录及关联寄出记录
        table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'
        db.execute(f"DELETE FROM {table} WHERE id=?", (item_id,))
        # 删除关联的寄出记录
        db.execute("DELETE FROM seal_send_records WHERE sample_id=?", (item_id,))
    else:
        # 恢复状态
        table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'
        db.execute(f"UPDATE {table} SET 状态='normal', updated_at=CURRENT_TIMESTAMP WHERE id=?", (item_id,))
    db.commit()
    return jsonify({'success': True, 'message': '已永久删除' if is_permanent else '已恢复'})

@app.route('/api/scrap', methods=['GET'])
@token_required
def list_scrap():
    db = get_db()
    
    sql = """SELECT ss.*, 
             CASE WHEN ss.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS 名称,
             CASE WHEN ss.item_type='seal' THEN s.序号 ELSE c.序号 END AS 序号
             FROM seal_scrapped_samples ss 
             LEFT JOIN seal_samples s ON ss.item_type='seal' AND ss.item_id=s.id 
             LEFT JOIN seal_color_samples c ON ss.item_type='color' AND ss.item_id=c.id 
             WHERE 1=1"""
    params = []
    
    # 动态筛选（报废记录的筛选字段来自ss表）
    i = 0
    while f'f_field_{i}' in request.args:
        field = request.args[f'f_field_{i}']
        op = request.args.get(f'f_op_{i}', 'contains')
        value = request.args.get(f'f_val_{i}', '')
        if not field or not value or not op:
            i += 1; continue
        
        # 特殊处理：item_type → ss.item_type, 名称→别名字段
        if field == 'item_type':
            sql += " AND ss.item_type=?"; params.append(value)
            i += 1; continue
        elif field == '名称':
            sql += " AND (ss.报废原因 LIKE ? OR ss.报废类型 LIKE ?)"
            params.extend(['%' + value + '%'] * 2)
            i += 1; continue
            
        safe_field = ''.join(c for c in field if c.isalnum() or c in ('_', 'Δ', '-'))
        if not safe_field: i += 1; continue
        
        op_sql_map = {
            'contains': ' LIKE ?', 'not_contains': ' NOT LIKE ?',
            'equals': ' = ?', 'not_equals': ' != ?',
            'gt': ' > ?', 'gte': ' >= ?', 'lt': ' < ?', 'lte': ' <= ?',
            'before': ' < ?', 'after': ' > ?'
        }
        sql_op = op_sql_map.get(op, ' LIKE ?')
        
        # 报废记录的字段都在ss表上，加前缀
        if safe_field in ('报废原因','报废类型','报废日期','报废审批人'):
            sql += f" AND ss.{safe_field}{sql_op}"
            params.append('%' + value + '%' if 'LIKE' in sql_op else value)
        else:
            sql += f" AND {safe_field}{sql_op}"
            params.append('%' + value + '%' if 'LIKE' in sql_op else value)
        i += 1
    
    # 兼容旧参数
    item_type = request.args.get('type', '')
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    keyword = request.args.get('keyword', '')
    if item_type: sql += " AND ss.item_type=?"; params.append(item_type)
    if start_date: sql += " AND ss.报费日期>=?"; params.append(start_date)  
    if end_date: sql += " AND ss.报费日期<=?"; params.append(end_date)
    if keyword: sql += " AND (ss.报废原因 LIKE ? OR ss.报废类型 LIKE ?)"; params.extend(['%'+keyword+'%']*2)

    sql += " ORDER BY ss.id DESC"
    rows = db.execute(sql, params).fetchall()
    results = [dict(r) for r in rows]
    return jsonify({'success': True, 'data': results})


# ==================== API: 管理员功能 ====================

@app.route('/api/admin/users', methods=['GET'])
@token_required
@admin_required
def admin_list_users():
    from datetime import datetime, timedelta
    db = get_db()
    search = request.args.get('search', '')
    if search:
        rows = db.execute("SELECT u.*, ic.code FROM users u LEFT JOIN seal_invitation_codes ic ON u.invitation_code_id=ic.id WHERE u.username LIKE ? OR u.real_name LIKE ?",
                         ('%'+search+'%', '%'+search+'%')).fetchall()
    else:
        rows = db.execute("SELECT u.*, ic.code FROM users u LEFT JOIN seal_invitation_codes ic ON u.invitation_code_id=ic.id ORDER BY u.id").fetchall()
    results = []
    for r in rows:
        d = dict(r)
        # 5分钟内有活动则视为在线 (last_active存储的是UTC时间，需转为本地时间)
        la = d.get('last_active')
        d['is_online'] = False
        if la:
            try:
                dt = datetime.strptime(str(la), '%Y-%m-%d %H:%M:%S')
                # 将UTC时间转为本地时间(UTC+8)后与当前时间比较
                dt_local = dt + timedelta(hours=8)
                d['is_online'] = (datetime.now() - dt_local) < timedelta(minutes=5)
            except: pass
        results.append(d)
    return jsonify({'success': True, 'data': results})

@app.route('/api/admin/users/<int:user_id>/status', methods=['PUT'])
@token_required
@admin_required
def admin_toggle_user(user_id):
    data = request.get_json()
    if user_id == g.current_user['id']:
        return jsonify({'success': False, 'message': '不能操作自己的账户'}), 400
    db = get_db()
    target = db.execute("SELECT is_admin FROM users WHERE id=?", (user_id,)).fetchone()
    if not target:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    if dict(target).get('is_admin'):
        return jsonify({'success': False, 'message': '不能禁用管理员账户'}), 400
    is_active = 1 if data.get('isActive', True) else 0
    db.execute("UPDATE users SET is_active=? WHERE id=?", (is_active, user_id))
    # 停用时立即清除在线状态
    if not is_active:
        db.execute("UPDATE users SET last_active=NULL WHERE id=?", (user_id,))
    db.commit()
    action = '启用' if is_active else '禁用'
    return jsonify({'success': True, 'message': f'用户已{action}'})

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@token_required
@admin_required
def admin_delete_user(user_id):
    if user_id == g.current_user['id']:
        return jsonify({'success': False, 'message': '不能删除自己的账户'}), 400
    db = get_db()
    target = db.execute("SELECT is_admin, invitation_code_id FROM users WHERE id=?", (user_id,)).fetchone()
    if not target:
        return jsonify({'success': False, 'message': '用户不存在'}), 404
    t = dict(target)
    if t.get('is_admin'):
        return jsonify({'success': False, 'message': '不能删除管理员账户'}), 400
    # 删除用户时，同步删除其关联的邀请码（如果有）
    invite_code_id = t.get('invitation_code_id')
    if invite_code_id:
        db.execute("DELETE FROM seal_invitation_codes WHERE id=?", (invite_code_id,))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    return jsonify({'success': True, 'message': '用户已注销'})

# 邀请码管理

@app.route('/api/admin/invitations', methods=['GET'])
@token_required
@admin_required
def list_invitations():
    db = get_db()
    rows = db.execute("SELECT ic.*, u.real_name as creator_name FROM seal_invitation_codes ic LEFT JOIN users u ON ic.created_by=u.id ORDER BY ic.id DESC").fetchall()
    return jsonify({'success': True, 'data': [dict(r) for r in rows]})

@app.route('/api/admin/invitations', methods=['POST'])
@token_required
@admin_required
def create_invitation():
    data = request.get_json()
    max_uses = int(data.get('maxUses', 1))
    expires_at = data.get('expiresAt', '')
    note = data.get('note', '')
    code = generate_invitation_code()
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""INSERT INTO seal_invitation_codes (code, note, max_uses, expires_at, is_active, created_by)
                      VALUES (?, ?, ?, ?, ?, ?)""", (code, note, max_uses, expires_at or None, 1, g.current_user['id']))
    db.commit()
    return jsonify({'success': True, 'message': '邀请码生成成功', 'data': {'code': code, 'id': cursor.lastrowid}})

@app.route('/api/admin/invitations/<int:code_id>', methods=['PUT'])
@token_required
@admin_required
def toggle_invitation(code_id):
    data = request.get_json()
    db = get_db()
    is_active = 1 if data.get('isActive', True) else 0
    db.execute("UPDATE seal_invitation_codes SET is_active=? WHERE id=?", (is_active, code_id))
    db.commit()
    action = '启用' if is_active else '停用'
    return jsonify({'success': True, 'message': f'邀请码已{action}'})

@app.route('/api/admin/invitations/<int:code_id>', methods=['DELETE'])
@token_required
@admin_required
def delete_invitation(code_id):
    db = get_db()
    # 检查是否已被使用（使用过的邀请码禁止删除）
    inv = db.execute("SELECT used_count FROM seal_invitation_codes WHERE id=?", (code_id,)).fetchone()
    if not inv:
        return jsonify({'success': False, 'message': '邀请码不存在'}), 404
    if dict(inv).get('used_count', 0) > 0:
        return jsonify({'success': False, 'message': '该邀请码已被使用，不允许删除'}), 400
    db.execute("DELETE FROM seal_invitation_codes WHERE id=?", (code_id,))
    db.commit()
    return jsonify({'success': True, 'message': '邀请码已删除'})

# 系统设置

@app.route('/api/admin/settings', methods=['GET'])
@token_required
@admin_required
def get_settings():
    db = get_db()
    rows = db.execute("SELECT * FROM seal_system_settings").fetchall()
    settings = {}
    for r in rows:
        settings[dict(r)['key']] = {'value': dict(r)['value'], 'description': dict(r)['description']}
    return jsonify({'success': True, 'data': settings})

@app.route('/api/admin/settings', methods=['PUT'])
@token_required
@admin_required
def update_settings():
    data = request.get_json()
    settings_data = data.get('settings', [])
    db = get_db()
    for s in settings_data:
        key = s.get('key')
        value = s.get('value')
        if key and value is not None:
            db.execute("UPDATE seal_system_settings SET value=?, updated_by=?, updated_at=CURRENT_TIMESTAMP WHERE key=?",
                       (value, g.current_user['id'], key))
    db.commit()
    return jsonify({'success': True, 'message': '系统设置保存成功'})


# ==================== API: 用户表格配置(筛选+列设置) ====================

@app.route('/api/table-configs/<page_key>', methods=['GET'])
@token_required
def get_table_config(page_key):
    """获取用户某页面的筛选/列配置"""
    db = get_db()
    config_type = request.args.get('type', 'columns')  # 'filter' or 'columns'
    row = db.execute(
        "SELECT config_data FROM seal_user_table_configs WHERE user_id=? AND page_key=? AND config_type=?",
        (g.current_user['id'], page_key, config_type)
    ).fetchone()
    if row:
        try:
            return jsonify({'success': True, 'data': json.loads(row['config_data']) if row['config_data'] else {}})
        except (json.JSONDecodeError, TypeError):
            return jsonify({'success': True, 'data': {}})
    return jsonify({'success': True, 'data': {}})

@app.route('/api/table-configs/<page_key>', methods=['PUT'])
@token_required
def save_table_config(page_key):
    """保存用户某页面的筛选/列配置"""
    data = request.get_json()
    config_type = data.get('type', 'columns')  # 'filter' or 'columns'
    config_data = data.get('config', {})
    db = get_db()
    existing = db.execute(
        "SELECT id FROM seal_user_table_configs WHERE user_id=? AND page_key=? AND config_type=?",
        (g.current_user['id'], page_key, config_type)
    ).fetchone()
    import json
    json_str = json.dumps(config_data, ensure_ascii=False) if isinstance(config_data, (dict, list)) else str(config_data)
    if existing:
        db.execute(
            "UPDATE seal_user_table_configs SET config_data=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (json_str, existing['id'])
        )
    else:
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO seal_user_table_configs (user_id, page_key, config_type, config_data) VALUES (?, ?, ?, ?)",
            (g.current_user['id'], page_key, config_type, json_str)
        )
    db.commit()
    return jsonify({'success': True, 'message': '配置已保存'})


# ==================== API: 处置记录（评定+报废统一视图） ====================

@app.route('/api/disposal-records', methods=['GET'])
@token_required
def list_disposal_records():
    """返回统一的处置记录列表（评定记录 + 报废记录合并，按时间倒序）"""
    db = get_db()
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    record_type = request_args_get = request.args.get('recordType', '')  # evaluation / scrap / all
    item_type = request.args.get('itemType', '')      # seal / color / all
    include_deleted = request.args.get('includeDeleted', '') == '1'  # 是否包含已删除

    # 收集动态筛选项
    dynamic_filters = []
    i = 0
    while f'f_field_{i}' in request.args:
        field = request.args[f'f_field_{i}']
        op = request_args_get = request.args.get(f'f_op_{i}', 'contains')
        value = request.args.get(f'f_val_{i}', '')
        if field and op:
            safe_field = ''.join(c for c in field if c.isalnum() or c in ('_', 'Δ', '-'))
            if safe_field:
                dynamic_filters.append({'field': safe_field, 'op': op, 'value': value})
        i += 1

    # 从动态筛选中提取 record_type（如果有），用于控制子查询范围
    rt_filter = None
    _other_filters = []
    for f in dynamic_filters:
        if f['field'] == 'record_type' and f['value']:
            rt_filter = f['value']
        else:
            _other_filters.append(f)

    # 如果没有通过动态筛选指定 record_type，则使用 URL 参数或默认全部
    if not rt_filter:
        rt_filter = request_args_get = request.args.get('recordType', '') or ''

    def apply_filters_to_sql(sql, params, table_prefix, filters):
        """将动态筛选条件追加到SQL，只处理与当前表相关的字段"""
        for f in filters:
            fname = f['field']
            fop = f['op']
            fval = f['value']

            # 空值=用户未选择/选了"全部"，跳过
            if not fval or fval.strip() == '':
                continue

            op_sql_map = {
                'contains': ' LIKE ?', 'not_contains': ' NOT LIKE ?',
                'equals': ' = ?', 'not_equals': ' != ?',
                'gt': ' > ?', 'gte': ' >= ?', 'lt': ' < ?', 'lte': ' <= ?',
                'before': ' < ?', 'after': ' > ?'
            }
            sql_op = op_sql_map.get(fop, ' LIKE ?')

            if fname == 'record_type':
                continue  # 已由外层 rt_filter 处理

            # 字段映射：前端字段名 → 实际数据库列名（带表前缀）
            col_mapping = {
                # 通用字段
                'item_type': f'{table_prefix}.item_type',
                # 评定记录特有
                '评定结果': f'{table_prefix}.评定结果',
                '评定人': f'{table_prefix}.评定人',
                '评定日期': f'{table_prefix}.评定日期',
                '计算ΔE值': f'{table_prefix}.计算ΔE值',
                '新有效期截止日': f'{table_prefix}.新有效期截止日',
                '评定说明': f'{table_prefix}.评定说明',
                # 报废记录特有
                '报废原因': f'{table_prefix}.报废原因',
                '报废类型': f'{table_prefix}.报废类型',
                '报废日期': f'{table_prefix}.报废日期',
                '报废审批人': f'{table_prefix}.报废审批人',
                '备注': f'{table_prefix}.备注',
                # JOIN出来的公共字段
                'item_name': None,
                'item_serial': None,
                '序号': None,
                '名称': None,
            }

            actual_col = col_mapping.get(fname)
            if actual_col is None:
                continue  # 字段不适用于此表，跳过

            if fname in ('item_name', '名称'):
                if 'LIKE' in sql_op:
                    sql += f" AND (s.封样件名称{sql_op} OR c.颜色名称{sql_op})"
                    params.extend(['%' + fval + '%'] * 2)
                else:
                    sql += f" AND (s.封样件名称{sql_op} OR c.颜色名称{sql_op})"
                    params.extend([fval] * 2)
            elif fname in ('item_serial', '序号'):
                if 'LIKE' in sql_op:
                    sql += f" AND (s.序号{sql_op} OR c.序号{sql_op})"
                    params.extend(['%' + fval + '%'] * 2)
                else:
                    sql += f" AND (s.序号{sql_op} OR c.序号{sql_op})"
                    params.extend([fval] * 2)
            else:
                sql += f" AND {actual_col}{sql_op}"
                if 'LIKE' in sql_op:
                    params.append('%' + fval + '%')
                else:
                    params.append(fval)

        return sql, params

    records = []

    # 查询评定记录
    if rt_filter in ('', 'evaluation', 'all'):
        del_cond = '' if include_deleted else " AND er.is_deleted=0"
        eval_sql = f"""SELECT er.*, 
            CASE WHEN er.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
            CASE WHEN er.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial,
            CASE WHEN er.item_type='seal' THEN s.有效期 ELSE c.有效期 END AS item_expiry
            FROM seal_evaluation_records er 
            LEFT JOIN seal_samples s ON er.item_type='seal' AND er.item_id=s.id 
            LEFT JOIN seal_color_samples c ON er.item_type='color' AND er.item_id=c.id 
            WHERE 1=1{del_cond}"""
        eval_params = []
        if item_type:
            eval_sql += " AND er.item_type=?"
            eval_params.append(item_type)
        eval_sql, eval_params = apply_filters_to_sql(eval_sql, eval_params, 'er', _other_filters)
        eval_sql += " ORDER BY er.id DESC"
        eval_rows = db.execute(eval_sql, eval_params).fetchall()
        for r in eval_rows:
            d = dict(r)
            d['record_type'] = 'evaluation'
            d['record_type_label'] = '评定记录'
            records.append(d)

    # 查询报废记录
    if rt_filter in ('', 'scrap', 'all'):
        del_cond = '' if include_deleted else " AND ss.is_deleted=0"
        scrap_sql = f"""SELECT ss.*, 
            CASE WHEN ss.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
            CASE WHEN ss.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial
            FROM seal_scrapped_samples ss 
            LEFT JOIN seal_samples s ON ss.item_type='seal' AND ss.item_id=s.id 
            LEFT JOIN seal_color_samples c ON ss.item_type='color' AND ss.item_id=c.id 
            WHERE 1=1{del_cond}"""
        scrap_params = []
        if item_type:
            scrap_sql += " AND ss.item_type=?"
            scrap_params.append(item_type)
        scrap_sql, scrap_params = apply_filters_to_sql(scrap_sql, scrap_params, 'ss', _other_filters)
        scrap_sql += " ORDER BY ss.id DESC"
        scrap_rows = db.execute(scrap_sql, scrap_params).fetchall()
        for r in scrap_rows:
            d = dict(r)
            d['record_type'] = 'scrap'
            d['record_type_label'] = '报废记录'
            records.append(d)

    # 合并并按 created_at 倒序
    records.sort(key=lambda x: x.get('created_at') or '', reverse=True)

    # 客户端分页
    total = len(records)
    start = (page - 1) * page_size
    paged = records[start:start + page_size]

    return jsonify({
        'success': True,
        'data': {
            'items': paged,
            'total': total,
            'page': page,
            'page_size': page_size,
            'total_pages': (total + page_size - 1) // page_size
        }
    })

@app.route('/api/disposal-records/export', methods=['GET'])
@token_required
def export_disposal_records():
    """导出处置记录为Excel"""
    db = get_db()
    records = []

    # 评定记录
    eval_rows = db.execute("""SELECT er.*, 
        CASE WHEN er.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
        CASE WHEN er.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial
        FROM seal_evaluation_records er 
        LEFT JOIN seal_samples s ON er.item_type='seal' AND er.item_id=s.id 
        LEFT JOIN seal_color_samples c ON er.item_type='color' AND er.item_id=c.id 
        ORDER BY er.id DESC""").fetchall()
    for r in eval_rows:
        d = dict(r)
        d['record_type'] = '评定记录'
        records.append(d)

    # 报废记录
    scrap_rows = db.execute("""SELECT ss.*, 
        CASE WHEN ss.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
        CASE WHEN ss.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial
        FROM seal_scrapped_samples ss 
        LEFT JOIN seal_samples s ON ss.item_type='seal' AND ss.item_id=s.id 
        LEFT JOIN seal_color_samples c ON ss.item_type='color' AND ss.item_id=c.id 
        ORDER BY ss.id DESC""").fetchall()
    for r in scrap_rows:
        d = dict(r)
        d['record_type'] = '报废记录'
        records.append(d)

    # 按 created_at 排序
    records.sort(key=lambda x: x.get('created_at') or '', reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "处置记录"

    headers = ['记录类型','对象类型','编号','名称','评定结果/报废原因','报废类型',
               '评定人/审批人','评定日期/报废日期','ΔE值','新有效期截止日',
               '评定说明/备注','创建时间']
    ws.append(headers)

    status_map = {'pass': '合格续期', 'fail': '不合格(报废)'}

    for r in records:
        rt = r.get('record_type', '')
        if rt == '评定记录':
            ws.append([
                '评定',
                '封样件' if r.get('item_type') == 'seal' else '色板',
                r.get('item_serial') or '-',
                r.get('item_name') or '-',
                status_map.get(r.get('评定结果', ''), r.get('评定结果', '')),
                '-',
                r.get('评定人', '-') or '-',
                r.get('评定日期', '-') or '-',
                r.get('计算ΔE值') if r.get('计算ΔE值') else '-',
                r.get('新有效期截止日') or '-',
                r.get('评定说明') or '-',
                r.get('created_at') or '-'
            ])
        else:
            ws.append([
                '报废',
                '封样件' if r.get('item_type') == 'seal' else '色板',
                r.get('item_serial') or '-',
                r.get('item_name') or '-',
                r.get('报废原因') or '-',
                r.get('报废类型') or '-',
                r.get('报废审批人', '-') or '-',
                r.get('报废日期') or '-',
                '-', '-', r.get('备注') or '-',
                r.get('created_at') or '-'
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='处置记录.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ==================== API: 处置记录 - 管理操作（软删除/恢复/永久删除） ====================

def _rollback_disposal_effect(record_type, rec):
    """根据处置记录类型，回滚台账数据（有效期或状态）"""
    db = get_db()
    rec = dict(rec)
    item_type = rec.get('item_type')
    item_id = rec.get('item_id')
    if not item_type or not item_id:
        return

    main_table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'

    if record_type == 'evaluation':
        # 评定记录：回滚有效期
        old_expiry = rec.get('旧有效期') or ''
        new_exp = rec.get('新有效期截止日') or ''
        if old_expiry and new_exp:
            db.execute(f"UPDATE {main_table} SET 有效期=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (old_expiry, item_id))
    elif record_type == 'scrap':
        # 报废记录：回滚状态
        old_status = rec.get('旧状态') or 'normal'
        db.execute(f"UPDATE {main_table} SET 状态=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                   (old_status, item_id))

def _reapply_disposal_effect(record_type, rec):
    """恢复处置记录时，重新应用其对台账的影响"""
    db = get_db()
    rec = dict(rec)
    item_type = rec.get('item_type')
    item_id = rec.get('item_id')
    if not item_type or not item_id:
        return

    main_table = 'seal_samples' if item_type == 'seal' else 'seal_color_samples'

    if record_type == 'evaluation':
        # 重新设置新有效期
        new_exp = rec.get('新有效期截止日') or ''
        result = rec.get('评定结果', '')
        if result == 'pass' and new_exp:
            db.execute(f"UPDATE {main_table} SET 有效期=?, 状态='normal', updated_at=CURRENT_TIMESTAMP WHERE id=?",
                       (new_exp, item_id))
    elif record_type == 'scrap':
        # 重新标记报废
        db.execute(f"UPDATE {main_table} SET 状态='scrapped', updated_at=CURRENT_TIMESTAMP WHERE id=?",
                   (item_id,))


@app.route('/api/disposal-records/<record_type>/<int:record_id>/delete', methods=['POST'])
@token_required
@admin_required
def soft_delete_disposal_record(record_type, record_id):
    """管理员：软删除处置记录（标记为已删除 + 回滚台账）"""
    db = get_db()
    if record_type == 'evaluation':
        table = 'seal_evaluation_records'
        label = '评定记录'
    elif record_type == 'scrap':
        table = 'seal_scrapped_samples'
        label = '报废记录'
    else:
        return jsonify({'success': False, 'message': '无效的记录类型'}), 400

    row = db.execute(f"SELECT * FROM {table} WHERE id=?", (record_id,)).fetchone()
    if not row:
        return jsonify({'success': False, 'message': f'{label}不存在'}), 404
    if dict(row).get('is_deleted'):
        return jsonify({'success': False, 'message': f'该{label}已被删除'}), 400

    # 回滚台账影响（恢复旧有效期 / 恢复旧状态）
    _rollback_disposal_effect(record_type, row)

    # 标记为已删除
    db.execute(f"UPDATE {table} SET is_deleted=1 WHERE id=?", (record_id,))
    db.commit()
    return jsonify({'success': True, 'message': f'{label}已移至回收站，台账数据已回滚'})


@app.route('/api/disposal-records/<record_type>/<int:record_id>/restore', methods=['POST'])
@token_required
@admin_required
def restore_disposal_record(record_type, record_id):
    """管理员：恢复已删除的处置记录（重新应用台账变更）"""
    db = get_db()
    if record_type == 'evaluation':
        table = 'seal_evaluation_records'
        label = '评定记录'
    elif record_type == 'scrap':
        table = 'seal_scrapped_samples'
        label = '报废记录'
    else:
        return jsonify({'success': False, 'message': '无效的记录类型'}), 400

    row = db.execute(f"SELECT * FROM {table} WHERE id=? AND is_deleted=1", (record_id,)).fetchone()
    if not row:
        return jsonify({'success': False, 'message': f'未找到待恢复的{label}'}), 404

    # 重新应用台账影响（设置新有效期 / 标记报废）
    _reapply_disposal_effect(record_type, row)

    db.execute(f"UPDATE {table} SET is_deleted=0 WHERE id=?", (record_id,))
    db.commit()
    return jsonify({'success': True, 'message': f'{label}已恢复，台账数据已更新'})


@app.route('/api/disposal-records/<record_type>/<int:record_id>/permanent-delete', methods=['POST'])
@token_required
@admin_required
def permanent_delete_disposal_record(record_type, record_id):
    """管理员：永久删除处置记录（不可恢复，回滚台账 + 清理关联寄出记录）"""
    db = get_db()
    if record_type == 'evaluation':
        table = 'seal_evaluation_records'
        label = '评定记录'
    elif record_type == 'scrap':
        table = 'seal_scrapped_samples'
        label = '报废记录'
    else:
        return jsonify({'success': False, 'message': '无效的记录类型'}), 400

    row = db.execute(f"SELECT * FROM {table} WHERE id=?", (record_id,)).fetchone()
    if not row:
        return jsonify({'success': False, 'message': f'{label}不存在'}), 404

    rec = dict(row)

    # 回滚台账影响
    _rollback_disposal_effect(record_type, row)

    # 永久删除处置记录本身
    db.execute(f"DELETE FROM {table} WHERE id=?", (record_id,))

    # 同时删除关联的寄出记录（sample_id 对应的封样件/色板）
    item_id = rec.get('item_id')
    if item_id and rec.get('item_type'):
        db.execute("DELETE FROM seal_send_records WHERE sample_id=?", (item_id,))

    db.commit()
    return jsonify({
        'success': True,
        'message': f'{label}及关联寄出记录已永久删除，台账数据已回滚',
        'data': {'deleted_id': record_id}
    })


@app.route('/api/disposal-records/deleted-list', methods=['GET'])
@token_required
@admin_required
def list_deleted_disposal_records():
    """管理员查看回收站（已删除的处置记录）"""
    db = get_db()

    records = []

    # 已删评定记录
    eval_rows = db.execute("""SELECT er.*, 
        CASE WHEN er.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
        CASE WHEN er.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial
        FROM seal_evaluation_records er 
        LEFT JOIN seal_samples s ON er.item_type='seal' AND er.item_id=s.id 
        LEFT JOIN seal_color_samples c ON er.item_type='color' AND er.item_id=c.id 
        WHERE er.is_deleted=1 ORDER BY er.id DESC""").fetchall()
    for r in eval_rows:
        d = dict(r)
        d['record_type'] = 'evaluation'
        d['record_type_label'] = '评定记录'
        records.append(d)

    # 已删报废记录
    scrap_rows = db.execute("""SELECT ss.*, 
        CASE WHEN ss.item_type='seal' THEN s.封样件名称 ELSE c.颜色名称 END AS item_name,
        CASE WHEN ss.item_type='seal' THEN s.序号 ELSE c.序号 END AS item_serial
        FROM seal_scrapped_samples ss 
        LEFT JOIN seal_samples s ON ss.item_type='seal' AND ss.item_id=s.id 
        LEFT JOIN seal_color_samples c ON ss.item_type='color' AND ss.item_id=c.id 
        WHERE ss.is_deleted=1 ORDER BY ss.id DESC""").fetchall()
    for r in scrap_rows:
        d = dict(r)
        d['record_type'] = 'scrap'
        d['record_type_label'] = '报废记录'
        records.append(d)

    records.sort(key=lambda x: x.get('created_at') or '', reverse=True)

    return jsonify({'success': True, 'data': {'items': records, 'total': len(records)}})


# ==================== API: 仪表盘统计 ====================

@app.route('/api/dashboard/stats', methods=['GET'])
@token_required
def dashboard_stats():
    db = get_db()
    
    seal_total = db.execute("SELECT COUNT(*) as cnt FROM seal_samples").fetchone()['cnt']
    color_total = db.execute("SELECT COUNT(*) as cnt FROM seal_color_samples").fetchone()['cnt']
    scrapped = db.execute("SELECT COUNT(*) as cnt FROM seal_scrapped_samples").fetchone()['cnt']
    
    # 待评定数：有效期≤30天（含已过期）且未报废
    today = get_china_date()
    pending = 0
    for t, tbl in [('seal', 'seal_samples'), ('color', 'seal_color_samples')]:
        rows = db.execute(f"""SELECT id, 有效期 FROM {tbl} WHERE 状态!='scrapped'""").fetchall()
        for r in rows:
            exp = r['有效期']
            if exp:
                try:
                    days_left = (datetime.strptime(str(exp)[:10], '%Y-%m-%d').date() - datetime.now().date()).days
                    if days_left <= 30: pending += 1  # 含已过期(负数)和未来30天
                except: pass
    
    return jsonify({
        'success': True,
        'data': {
            'sealTotal': seal_total,
            'colorTotal': color_total,
            'pendingEval': pending,
            'scrappedTotal': scrapped
        }
    })

@app.route('/api/dashboard/warnings', methods=['GET'])
@token_required
def dashboard_warnings():
    db = get_db()
    warnings = []
    today = datetime.now().date()
    
    # 封样件预警（含已过期）
    seal_rows = db.execute("SELECT id, 'seal' as type, 序号, 项目, 封样件名称 as name, 有效期 FROM seal_samples WHERE 状态!='scrapped' AND 有效期 IS NOT NULL").fetchall()
    for r in seal_rows:
        d = dict(r)
        try:
            exp = datetime.strptime(str(d['有效期'])[:10], '%Y-%m-%d').date()
            days_left = (exp - today).days
            if days_left <= 30:  # 含已过期
                warnings.append({
                    'id': d['id'], 'type': 'seal', 'name': d['name'] or f"#{d['序号']}",
                    'expiry': d['有效期'], 'daysLeft': days_left,
                    'status': 'expired' if days_left < 0 else ('pending_eval' if days_left <= 30 else 'normal')
                })
        except: pass
    
    # 色板预警（含已过期）
    color_rows = db.execute("SELECT id, 'color' as type, 序号, 客户, 颜色名称 as name, 有效期 FROM seal_color_samples WHERE 状态!='scrapped' AND 有效期 IS NOT NULL").fetchall()
    for r in color_rows:
        d = dict(r)
        try:
            exp = datetime.strptime(str(d['有效期'])[:10], '%Y-%m-%d').date()
            days_left = (exp - today).days
            if days_left <= 30:  # 含已过期
                warnings.append({
                    'id': d['id'], 'type': 'color', 'name': d['name'] or f"#{d['序号']} ({d['客户']})",
                    'expiry': d['有效期'], 'daysLeft': days_left,
                    'status': 'expired' if days_left < 0 else ('pending_eval' if days_left <= 30 else 'normal')
                })
        except: pass
    
    warnings.sort(key=lambda x: x['daysLeft'])
    return jsonify({'success': True, 'data': warnings[:10]})


# ==================== 页面路由 ====================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/templates/<path:filename>')
def serve_sub_page(filename):
    """Serve sub-page templates for AJAX loading (index.html fetches these)"""
    template_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
    if not filename.endswith('.html'):
        return jsonify({'success': False, 'message': 'Not allowed'}), 403
    try:
        return send_from_directory(template_dir, filename)
    except:
        return jsonify({'success': False, 'message': 'Page not found'}), 404

@app.route('/favicon.ico')
def favicon():
    """Return a minimal SVG favicon"""
    svg = '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 32 32"><rect width="32" height="32" rx="6" fill="#1a73e8"/><text x="16" y="22" text-anchor="middle" fill="white" font-size="18" font-family="Arial" font-weight="bold">封</text></svg>'
    return svg, 200, {'Content-Type': 'image/svg+xml'}


# ==================== 启动入口 ====================

if __name__ == '__main__':
    print("="*50)
    print("  封样件及色板接收登记管理系统")
    print("  正在初始化数据库...")
    init_db()
    print("  数据库初始化完成")
    print(f"  服务启动地址: http://127.0.0.1:5000")
    print(f"  登录账号: admin / admin")
    print("="*50)
    app.run(host='0.0.0.0', port=5000, debug=True)
